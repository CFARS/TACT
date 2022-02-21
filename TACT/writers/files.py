import pandas as pd
import numpy as np
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import argparse
import re
from sklearn import linear_model
from sklearn.metrics import mean_squared_error, r2_score
import warnings
import os
import math
import datetime
from future.utils import itervalues, iteritems
import json
import requests
from glob2 import glob
import matplotlib.pyplot as plt
from string import printable


def configure_for_printing(df, Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass = False, byTIRefbin = False):

    if isinstance(df, list) == False:
        pass
    else:
        for val in df:
            for i in val:
                if isinstance(i, pd.DataFrame) == False:
                    pass
                else:
                    try:
                        dat = i[i.columns.to_list()[0][0]]
                        stats = list(set([tup[0] for tup in dat.columns.to_list()]))
                        for s in stats:
                            try:
                               new_data = dat[s].add_prefix(str(s + '_'))
                               if stabilityClass:
                                   new_index = [str('stability_' + stabilityClass + '_' + n)
                                                for n in new_data.index.to_list()]
                                   new_data.index = new_index
                               if s == 'mean' and new_data.columns.name == 'bins':
                                   Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                               elif s == 'mean' and new_data.columns.name == 'bins_p5':
                                   Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])
                               elif s == 'std' and new_data.columns.name == 'bins':
                                   Result_df_std_1mps = pd.concat([Result_df_std_1mps,new_data])
                               elif s == 'std' and new_data.columns.name == 'bins_p5':
                                   Result_df_std_05mps = pd.concat([Result_df_std_05mps,new_data])
                            except:
                               print (str('No data to write in one of the dataframes of results'))
                               rowNumber += 1
                    except:
                        pass
    if byTIRefbin:
        if isinstance(df,list) == False:
            pass
        else:
            for val in df:
                for i in val:
                    if isinstance(i, pd.DataFrame) == False:
                        pass
                    else:
                        try:
                            dat = i[i.columns.to_list()[0][0]]
                            stats = list(set([tup[0] for tup in dat.columns.to_list()]))
                            for s in stats:
                                try:
                                    new_data = dat[s].add_prefix(str(s + '_'))
                                    if stabilityClass:
                                        new_index = [str('stability_' + stabilityClass + '_' + n)
                                                     for n in new_data.index.to_list()]
                                        new_data.index = new_index
                                    if s == 'mean' and new_data.columns.name == 'RefTI_bins':
                                        Result_df_mean_tiRef = pd.concat([Result_df_mean_tiRef,new_data])
                                    elif s == 'std' and new_data.columns.name == 'RefTI_bins':
                                        Result_df_std_tiRef = pd.concat([Result_df_std_tiRef,new_data])
                                except:
                                   print (str('No data to write in one of the dataframes of results'))
                                   rowNumber += 1
                        except:
                           pass

    else:
        Result_df_mean_tiRef = Result_df_mean_tiRef
        Result_df_std_tiRef = Result_df_std_tiRef


    return Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps,Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef


def write_resultstofile(df, ws, r_start, c_start):
    # write the regression results to file.
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, r_start):
        for c_idx, value in enumerate(row, c_start):
            ws.cell(row=r_idx, column=c_idx, value=value)


def write_all_resultstofile(reg_results, baseResultsLists, count_1mps, count_05mps, count_1mps_train, count_05mps_train,
                            count_1mps_test, count_05mps_test, name_1mps_tke, name_1mps_alpha_Ane, name_1mps_alpha_RSD,
                            name_05mps_tke, name_05mps_alpha_Ane, name_05mps_alpha_RSD, count_05mps_tke, count_05mps_alpha_Ane, count_05mps_alpha_RSD,
                            count_1mps_tke, count_1mps_alpha_Ane, count_1mps_alpha_RSD, results_filename, siteMetadata, filterMetadata,
                            Timestamps, timestamp_train, timestamp_test, regimeBreakdown_tke, regimeBreakdown_ane, regimeBreakdown_rsd,
                            Ht_1_ane, Ht_2_ane, extrap_metadata, reg_results_class1, reg_results_class2, reg_results_class3,
                            reg_results_class4, reg_results_class5,reg_results_class1_alpha, reg_results_class2_alpha, reg_results_class3_alpha,
                            reg_results_class4_alpha, reg_results_class5_alpha, Ht_1_rsd, Ht_2_rsd, ResultsLists_stability, ResultsLists_stability_alpha_RSD,
                            ResultsLists_stability_alpha_Ane, stabilityFlag, cup_alphaFlag, RSD_alphaFlag, TimeTestA_baseline_df, TimeTestB_baseline_df, TimeTestC_baseline_df,TimeTestA_corrections_df,TimeTestB_corrections_df,TimeTestC_corrections_df):

    wb = Workbook()
    ws = wb.active

    Dist_stats_df = pd.DataFrame()

    # all baseline regressions
    # ------------------------
    a = wb.create_sheet(title='Baseline Results')
    rowNumber = 1
    write_resultstofile(reg_results,a,rowNumber,1)
    rowNumber += len(reg_results) + 3
    col = 1

    if stabilityFlag:
        write_resultstofile(reg_results_class1[0],a,rowNumber,col)
        rowNumber2 = rowNumber + len(reg_results) + 3
        write_resultstofile(reg_results_class2[0],a,rowNumber2,col)
        rowNumber3 = rowNumber2 + len(reg_results) + 3
        write_resultstofile(reg_results_class3[0],a,rowNumber3,col)
        rowNumber4 = rowNumber3 + len(reg_results) + 3
        write_resultstofile(reg_results_class4[0],a,rowNumber4,col)
        rowNumber5 = rowNumber4 + len(reg_results) + 3
        write_resultstofile(reg_results_class5[0],a,rowNumber5,col)

        for i in range(1,len(reg_results_class1)):
            col += reg_results_class1[0].shape[1] + 2
            write_resultstofile(reg_results_class1[i],a,rowNumber,col)
            write_resultstofile(reg_results_class2[i],a,rowNumber2,col)
            write_resultstofile(reg_results_class3[i],a,rowNumber3,col)
            write_resultstofile(reg_results_class4[i],a,rowNumber4,col)
            write_resultstofile(reg_results_class5[i],a,rowNumber5,col)

    rowNumber = rowNumber5 + len(reg_results) + 3

    if cup_alphaFlag:
        write_resultstofile(reg_results_class1_alpha['Ane'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class1_alpha['Ane']) + 3
        write_resultstofile(reg_results_class2_alpha['Ane'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class2_alpha['Ane']) + 3
        write_resultstofile(reg_results_class3_alpha['Ane'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class3_alpha['Ane']) + 3
        write_resultstofile(reg_results_class4_alpha['Ane'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class4_alpha['Ane']) + 3
        write_resultstofile(reg_results_class5_alpha['Ane'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class5_alpha['Ane']) + 3

    if RSD_alphaFlag:
        write_resultstofile(reg_results_class1_alpha['RSD'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class1_alpha['RSD']) + 3
        write_resultstofile(reg_results_class2_alpha['RSD'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class2_alpha['RSD']) + 3
        write_resultstofile(reg_results_class3_alpha['RSD'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class3_alpha['RSD']) + 3
        write_resultstofile(reg_results_class4_alpha['RSD'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class4_alpha['RSD']) + 3
        write_resultstofile(reg_results_class5_alpha['RSD'],a,rowNumber,1)
        rowNumber = rowNumber + len(reg_results_class5_alpha['RSD']) + 3


    # total bin counts and length of observations
    totalcount_1mps = count_1mps.sum().sum()
    totalcount_1mps_train = count_1mps_train.sum().sum()
    totalcount_1mps_test = count_1mps_test.sum().sum()
    totalcount_05mps = count_05mps.sum().sum()
    totalcount_05mps_train = count_05mps_train.sum().sum()
    totalcount_05mps_test = count_05mps_test.sum().sum()

    name_1mps_tke = []
    name_1mps_alpha_Ane = []
    name_1mps_alpha_RSD = []

    rowNumber = rowNumber + 2
    a.cell(row=rowNumber, column=1, value='Total Count (number of observations)')
    a.cell(row=rowNumber, column=2, value=totalcount_1mps)
    a.cell(row=rowNumber, column=4, value='Total Count in Training Subset (number of observations)')
    a.cell(row=rowNumber, column=5, value=totalcount_1mps_train)
    a.cell(row=rowNumber, column=7, value='Total Count in Training Subset (number of observations)')
    a.cell(row=rowNumber, column=8, value=totalcount_1mps_test)
    rowNumber += 2

    a.cell(row=rowNumber, column=1, value='Bin Counts')
    rowNumber += 1
    c_1mps = count_1mps['RSD_WS']['count']
    c_1mps.index=['count']
    write_resultstofile(c_1mps, a, rowNumber, 1)
    rowNumber += 4
    c_05mps = count_05mps['RSD_WS']['count']
    c_05mps.index=['count']
    write_resultstofile(c_05mps, a, rowNumber, 1)
    rowNumber += 4

    a.cell(row=rowNumber, column=1, value='Bin Counts (Train)')
    rowNumber += 1
    c_1mps_train = count_1mps_train['RSD_WS']['count']
    c_1mps_train.index=['count']
    write_resultstofile(c_1mps_train, a, rowNumber, 1)
    rowNumber += 4
    c_05mps_train = count_05mps_train['RSD_WS']['count']
    c_05mps_train.index=['count']
    write_resultstofile(c_05mps_train, a, rowNumber, 1)
    rowNumber += 4

    a.cell(row=rowNumber, column=1, value='Bin Counts (Test)')
    rowNumber += 1
    c_1mps_test = count_1mps_test['RSD_WS']['count']
    c_1mps_test.index=['count']
    write_resultstofile(c_1mps_test, a, rowNumber, 1)
    rowNumber += 4
    c_05mps_test = count_05mps_test['RSD_WS']['count']
    c_05mps_test.index=['count']
    write_resultstofile(c_05mps_test, a, rowNumber, 1)
    rowNumber += 4

    for c in range(0,len(count_1mps_tke)):
        a.cell(row=rowNumber, column=1, value=str('Bin Counts TKE' + str(c + 1)))
        rowNumber += 1
        try:
            c_1mps_test = count_1mps_tke[c]['RSD_WS']['count']
            c_1mps_test.index=['count']
        except:
            c_1mps_test = pd.DataFrame()
        write_resultstofile(c_1mps_test, a, rowNumber, 1)
        rowNumber += 4
        try:
            c_05mps_test = count_05mps_tke[c]['RSD_WS']['count']
            c_05mps_test.index=['count']
        except:
            c_05mps_test = pd.DataFrame()
        write_resultstofile(c_05mps_test, a, rowNumber, 1)
        rowNumber += 4

    for c in range(0,len(count_1mps_alpha_Ane)):
        a.cell(row=rowNumber, column=1, value=str('Bin Counts alpha Ane' + str(c + 1)))
        rowNumber += 1
        try:
            c_1mps_test = count_1mps_alpha_Ane[c]['RSD_WS']['count']
            c_1mps_test.index=['count']
        except:
            c_1mps_test = pd.DataFrame()
        write_resultstofile(c_1mps_test, a, rowNumber, 1)
        rowNumber += 4
        try:
            c_05mps_test = count_05mps_alpha_Ane[c]['RSD_WS']['count']
            c_05mps_test.index=['count']
        except:
            c_05mps_test = pd.DataFrame()
        write_resultstofile(c_05mps_test, a, rowNumber, 1)
        rowNumber += 4

    for c in range(0,len(count_1mps_alpha_RSD)):
        a.cell(row=rowNumber, column=1, value=str('Bin Counts alpha RSD' + str(c + 1)))
        rowNumber += 1
        try:
            c_1mps_test = count_1mps_alpha_RSD[c]['RSD_WS']['count']
            c_1mps_test.index=['count']
        except:
            c_1mps_test = pd.DataFrame()
        write_resultstofile(c_1mps_test, a, rowNumber, 1)
        rowNumber += 4
        try:
            c_05mps_test = count_05mps_alpha_RSD[c]['RSD_WS']['count']
            c_05mps_test.index=['count']
        except:
            c_05mps_test = pd.DataFrame()
        write_resultstofile(c_05mps_test, a, rowNumber, 1)
        rowNumber += 4

    totalcount_1mps_alpha_Ane = []
    totalcount_1mps_alpha_RSD = []

    totalcountTime_days = (totalcount_1mps * 10)/ 60/ 24
    a.cell(row=rowNumber, column=1, value='Total Time (days)')
    a.cell(row=rowNumber, column=2, value=totalcountTime_days)
    totalcountTime_days_train = (totalcount_1mps_train * 10)/60/24
    a.cell(row=rowNumber, column=4, value='Total Time Train (days)')
    a.cell(row=rowNumber, column=5, value=totalcountTime_days_train)
    totalcountTime_days_test = (totalcount_1mps_test * 10)/60/24
    a.cell(row=rowNumber, column=7, value='Total Time Test (days)')
    a.cell(row=rowNumber, column=8, value=totalcountTime_days_test)

    rowNumber += 2

    a.cell(row=rowNumber, column=1, value='Start Timestamp')
    a.cell(row=rowNumber, column=2, value=str(Timestamps[0]))
    a.cell(row=rowNumber, column=4, value='Start Timestamp (Train)')
    a.cell(row=rowNumber, column=5, value=str(timestamp_train[0]))
    a.cell(row=rowNumber, column=7, value='Start Timestamp (Test)')
    a.cell(row=rowNumber, column=8, value=str(timestamp_test[-1]))
    a.cell(row=rowNumber+1, column=1, value='End Timestamp')
    a.cell(row=rowNumber+1, column=2, value=str(Timestamps[-1]))
    a.cell(row=rowNumber+1, column=4, value='End Timestamp (Train)')
    a.cell(row=rowNumber+1, column=5, value=str(timestamp_train[0]))
    a.cell(row=rowNumber+1, column=7, value='End Timestamp (Test)')
    a.cell(row=rowNumber+1, column=8, value=str(timestamp_test[-1]))
    rowNumber +=3

    if stabilityFlag:
        a.cell(row=rowNumber, column=1, value='Stability TKE')
        rowNumber +=1
        write_resultstofile(regimeBreakdown_tke,a,rowNumber,1)
        rowNumber += 9
    if cup_alphaFlag:
        a.cell(row=rowNumber, column=1, value='Stability alpha: tower')
        rowNumber +=1
        a.cell(row=rowNumber, column=1, value='Heights for alpha calculation: tower')
        a.cell(row=rowNumber, column=2, value=Ht_1_ane)
        a.cell(row=rowNumber, column=3, value=Ht_2_ane)
        rowNumber += 2
        write_resultstofile(regimeBreakdown_ane,a, rowNumber,1)
        rowNumber +=8
    if RSD_alphaFlag:
        a.cell(row=rowNumber, column=1, value='Stability alpha: RSD')
        rowNumber +=1
        a.cell(row=rowNumber, column=1, value='Heights for alpha calculation: RSD')
        a.cell(row=rowNumber, column=2, value=Ht_1_rsd)
        a.cell(row=rowNumber, column=3, value=Ht_2_rsd)
        rowNumber += 2
        write_resultstofile(regimeBreakdown_rsd,a, rowNumber,1)
        rowNumber +=7

    # Metadata
    # --------
    b = wb.create_sheet(title='Metadata')
    rowNumber = 1
    b.cell(row=rowNumber, column=1, value='Software version: ')
    b.cell(row= rowNumber, column = 2, value = '1.1.0')
    b.cell(row= rowNumber, column=3, value = datetime.datetime.now())
    rowNumber += 3
    for r in dataframe_to_rows(siteMetadata, index=False):
        b.append(r)
        rowNumber = rowNumber + 1

    rowNumber += 2
    b.cell(row=rowNumber, column=1, value='Filter Metadata')
    rowNumber +=1
    write_resultstofile(filterMetadata,b,rowNumber,1)
    rowNumber += 2
    b.cell(row=rowNumber, column=1, value='Extrapolation Metadata')
    rowNumber +=1
    write_resultstofile(extrap_metadata,b,rowNumber,1)
    rowNumber +=9
    b.cell(row=rowNumber, column=1, value='Corrections Metadata')
    rowNumber +=1

    for c in baseResultsLists['correctionTagList_']:
         b.cell(row=rowNumber, column=1, value='Correction applied:')
         b.cell(row=rowNumber, column=2, value = c)
         rowNumber +=1

    # Time Sensitivity Tests
    # ----------------------
    # TimeTestA, TimeTestB, TimeTestC
    Ta = wb.create_sheet(title='Sensitivity2TestLengthA')
    rowNumber = 1
    Ta.cell(row=rowNumber, column = 1, value = 'baseline')
    rowNumber += 1
    write_resultstofile(TimeTestA_baseline_df, Ta, rowNumber,1)
    rowNumber += (len(TimeTestA_baseline_df)) + 4

    for key in TimeTestA_corrections_df:
        Ta.cell(row= rowNumber, column = 1, value = key)
        rowNumber += 1
        write_resultstofile(TimeTestA_corrections_df[key], Ta, rowNumber, 1)
        rowNumber += len(TimeTestA_corrections_df[key]) + 3

    Tb = wb.create_sheet(title='Sensitivity2TestLengthB')
    rowNumber = 1
    Tb.cell(row=rowNumber, column = 1, value = 'baseline')
    rowNumber += 1
    write_resultstofile(TimeTestB_baseline_df, Tb, rowNumber,1)
    rowNumber += (len(TimeTestB_baseline_df))

    for key in TimeTestB_corrections_df:
        Tb.cell(row= rowNumber, column = 1, value = key)
        rowNumber += 1
        write_resultstofile(TimeTestB_corrections_df[key], Tb, rowNumber, 1)
        rowNumber += len(TimeTestB_corrections_df[key]) + 3

    Tc = wb.create_sheet(title='Sensitivity2TestLengthC')
    rowNumber = 1
    Tc.cell(row=rowNumber, column = 1, value = 'baseline')
    rowNumber += 1
    write_resultstofile(TimeTestC_baseline_df, Tc, rowNumber,1)
    rowNumber += (len(TimeTestC_baseline_df))

    for key in TimeTestC_corrections_df:
        Tc.cell(row= rowNumber, column = 1, value = key)
        rowNumber += 1
        write_resultstofile(TimeTestC_corrections_df[key], Tc, rowNumber, 1)
        rowNumber += len(TimeTestC_corrections_df[key]) + 3

    # record results for each correction method
    # -----------------------------------------
    for correction in baseResultsLists['correctionTagList_']: # create tab for each correction method
        sheetName = correction

        for i in baseResultsLists['correctionTagList_']:
            if i == correction:
                idx = baseResultsLists['correctionTagList_'].index(i)

        TI_MBE_j_ = baseResultsLists['TI_MBEList_'][idx]
        TI_Diff_j_ = baseResultsLists['TI_DiffList_'][idx]
        TI_Diff_r_ = baseResultsLists['TI_DiffRefBinsList_'][idx]
        TI_RMSE_j_ = baseResultsLists['TI_RMSEList_'][idx]
        RepTI_MBE_j_ = baseResultsLists['RepTI_MBEList_'][idx]
        RepTI_Diff_j_ = baseResultsLists['RepTI_DiffList_'][idx]
        RepTI_Diff_r_ = baseResultsLists['RepTI_DiffRefBinsList_'][idx]
        RepTI_RMSE_j_ = baseResultsLists['RepTI_RMSEList_'][idx]
        rep_TI_results_1mps = baseResultsLists['rep_TI_results_1mps_List_'][idx]
        rep_TI_results_05mps = baseResultsLists['rep_TI_results_05mps_List_'][idx]
        TIbybin = baseResultsLists['TIBinList_'][idx]
        TIbyRefbin = baseResultsLists['TIRefBinList_'][idx]
        total_stats = baseResultsLists['total_StatsList_'][idx]
        belownominal_stats = baseResultsLists['belownominal_statsList_'][idx]
        abovenominal_stats = baseResultsLists['abovenominal_statsList_'][idx]
        lm_corr = baseResultsLists['lm_CorrList_'][idx]
        Dist_stats_df = pd.concat([Dist_stats_df,baseResultsLists['Distribution_statsList_'][idx]], axis=1)
        correctionTag = baseResultsLists['correctionTagList_'][idx]

        if stabilityFlag:
            TI_MBE_j_stability = ResultsLists_stability['TI_MBEList_stability_'][idx]
            TI_Diff_j_stability =  ResultsLists_stability['TI_DiffList_stability_'][idx]
            TI_Diff_r_stability = ResultsLists_stability['TI_DiffRefBinsList_stability_'][idx]
            TI_RMSE_j_stability =  ResultsLists_stability['TI_RMSEList_stability_'][idx]
            RepTI_MBE_j_stability = ResultsLists_stability['RepTI_MBEList_stability_'][idx]
            RepTI_Diff_j_stability = ResultsLists_stability['RepTI_DiffList_stability_'][idx]
            RepTI_Diff_r_stability = ResultsLists_stability['RepTI_DiffRefBinsList_stability_'][idx]
            RepTI_RMSE_j_stability = ResultsLists_stability['RepTI_RMSEList_stability_'][idx]
            rep_TI_results_1mps_stability = ResultsLists_stability['rep_TI_results_1mps_List_stability_'][idx]
            rep_TI_results_05mps_stability = ResultsLists_stability['rep_TI_results_05mps_List_stability_'][idx]
            TIbybin_stability = ResultsLists_stability['TIBinList_stability_'][idx]
            TIbyRefbin_stability = ResultsLists_stability['TIRefBinList_stability_'][idx]
            total_stats_stability = ResultsLists_stability['total_StatsList_stability_'][idx]
            belownominal_stats_stability = ResultsLists_stability['belownominal_statsList_stability_'][idx]
            abovenominal_stats_stability = ResultsLists_stability['abovenominal_statsList_stability_'][idx]
            lm_corr_stability = ResultsLists_stability['lm_CorrList_stability_'][idx]
            corrrectionTag_stability = ResultsLists_stability['correctionTagList_stability_'][idx]
            for i in ResultsLists_stability['Distribution_statsList_stability_'][idx]:
                if isinstance(i, pd.DataFrame):
                    Dist_stats_df = pd.concat([Dist_stats_df,i], axis=1)

        if cup_alphaFlag:
            TI_MBE_j_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['TI_MBEList_stability_alpha_Ane'][idx]
            TI_Diff_j_stability_alpha_Ane =  ResultsLists_stability_alpha_Ane['TI_DiffList_stability_alpha_Ane'][idx]
            TI_Diff_r_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['TI_DiffRefBinsList_stability_alpha_Ane'][idx]
            TI_RMSE_j_stability_alpha_Ane =  ResultsLists_stability_alpha_Ane['TI_RMSEList_stability_alpha_Ane'][idx]
            RepTI_MBE_j_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['RepTI_MBEList_stability_alpha_Ane'][idx]
            RepTI_Diff_j_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['RepTI_DiffList_stability_alpha_Ane'][idx]
            RepTI_Diff_r_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['RepTI_DiffRefBinsList_stability_alpha_Ane'][idx]
            RepTI_RMSE_j_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['RepTI_RMSEList_stability_alpha_Ane'][idx]
            rep_TI_results_1mps_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['rep_TI_results_1mps_List_stability_alpha_Ane'][idx]
            rep_TI_results_05mps_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['rep_TI_results_05mps_List_stability_alpha_Ane'][idx]
            TIbybin_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['TIBinList_stability_alpha_Ane'][idx]
            TIbyRefbin_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['TIRefBinList_stability_alpha_Ane'][idx]
            total_stats_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['total_StatsList_stability_alpha_Ane'][idx]
            belownominal_stats_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['belownominal_statsList_stability_alpha_Ane'][idx]
            abovenominal_stats_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['abovenominal_statsList_stability_alpha_Ane'][idx]
            lm_corr_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['lm_CorrList_stability_alpha_Ane'][idx]
            corrrectionTag_stability_alpha_Ane = ResultsLists_stability_alpha_Ane['correctionTagList_stability_alpha_Ane'][idx]
            for i in ResultsLists_stability_alpha_Ane['Distribution_statsList_stability_alpha_Ane'][idx]:
                if isinstance(i, pd.DataFrame):
                    Dist_stats_df = pd.concat([Dist_stats_df,i], axis=1)

        if RSD_alphaFlag:
            TI_MBE_j_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['TI_MBEList_stability_alpha_RSD'][idx]
            TI_Diff_j_stability_alpha_RSD =  ResultsLists_stability_alpha_RSD['TI_DiffList_stability_alpha_RSD'][idx]
            TI_Diff_r_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['TI_DiffRefBinsList_stability_alpha_RSD'][idx]
            TI_RMSE_j_stability_alpha_RSD =  ResultsLists_stability_alpha_RSD['TI_RMSEList_stability_alpha_RSD'][idx]
            RepTI_MBE_j_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['RepTI_MBEList_stability_alpha_RSD'][idx]
            RepTI_Diff_j_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['RepTI_DiffList_stability_alpha_RSD'][idx]
            RepTI_Diff_r_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['RepTI_DiffRefBinsList_stability_alpha_RSD'][idx]
            RepTI_RMSE_j_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['RepTI_RMSEList_stability_alpha_RSD'][idx]
            rep_TI_results_1mps_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['rep_TI_results_1mps_List_stability_alpha_RSD'][idx]
            rep_TI_results_05mps_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['rep_TI_results_05mps_List_stability_alpha_RSD'][idx]
            TIbybin_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['TIBinList_stability_alpha_RSD'][idx]
            TIbyRefbin_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['TIRefBinList_stability_alpha_RSD'][idx]
            total_stats_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['total_StatsList_stability_alpha_RSD'][idx]
            belownominal_stats_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['belownominal_statsList_stability_alpha_RSD'][idx]
            abovenominal_stats_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['abovenominal_statsList_stability_alpha_RSD'][idx]
            lm_corr_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['lm_CorrList_stability_alpha_RSD'][idx]
            corrrectionTag_stability_alpha_RSD = ResultsLists_stability_alpha_RSD['correctionTagList_stability_alpha_RSD'][idx]
            for i in ResultsLists_stability_alpha_RSD['Distribution_statsList_stability_alpha_RSD'][idx]:
                if isinstance(i, pd.DataFrame):
                    Dist_stats_df = pd.concat([Dist_stats_df,i], axis=1)

        ws = wb.create_sheet(title=sheetName)

        rowNumber = 1
        ws.cell(row=rowNumber, column=1, value='Corrected RSD Regression Results')
        ws.cell(row=rowNumber, column=2, value='m')
        ws.cell(row=rowNumber, column=3, value='c')
        ws.cell(row=rowNumber, column=4, value='r-squared')
        ws.cell(row=rowNumber, column=5, value='mean difference')
        ws.cell(row=rowNumber, column=6, value='mse')
        ws.cell(row=rowNumber, column=7, value='rmse')
        
        className = 1
        if stabilityFlag:
            for i in lm_corr_stability:
                 start = className*8 + 1
                 corrName = str('Corrected RSD Regression Results, stability subset (TKE)' + '_' + 'class_' + str(className))
                 ws.cell(row=rowNumber, column=start, value= corrName)
                 ws.cell(row=rowNumber, column=start+1, value='m')
                 ws.cell(row=rowNumber, column=start+2, value='c')
                 ws.cell(row=rowNumber, column=start+3, value='r-squared')
                 ws.cell(row=rowNumber, column=start+4, value='mean difference')
                 ws.cell(row=rowNumber, column=start+5, value='mse')
                 ws.cell(row=rowNumber, column=start+6, value='rmse')
                 className += 1

        className = 1
        if cup_alphaFlag:
             rowNumber = 13
             for i in lm_corr_stability_alpha_Ane:
                 start = className*8 + 1
                 corrName = str('Corrected RSD Regression Results, stability subset (cup alpha)' + '_' + 'class_' + str(className))
                 ws.cell(row=rowNumber, column=start, value= corrName)
                 ws.cell(row=rowNumber, column=start+1, value='m')
                 ws.cell(row=rowNumber, column=start+2, value='c')
                 ws.cell(row=rowNumber, column=start+3, value='r-squared')
                 ws.cell(row=rowNumber, column=start+4, value='mean difference')
                 ws.cell(row=rowNumber, column=start+5, value='mse')
                 ws.cell(row=rowNumber, column=start+6, value='rmse')
                 className += 1

        className = 1
        if RSD_alphaFlag:
             rowNumber = 25
             for i in lm_corr_stability_alpha_Ane:
                 start = className*8 + 1
                 corrName = str('Corrected RSD Regression Results, stability subset (RSD alpha)' + '_' + 'class_' + str(className))
                 ws.cell(row=rowNumber, column=start, value= corrName)
                 ws.cell(row=rowNumber, column=start+1, value='m')
                 ws.cell(row=rowNumber, column=start+2, value='c')
                 ws.cell(row=rowNumber, column=start+3, value='r-squared')
                 ws.cell(row=rowNumber, column=start+4, value='mean difference')
                 ws.cell(row=rowNumber, column=start+5, value='mse')
                 ws.cell(row=rowNumber, column=start+6, value='rmse')
                 className += 1

         # correction regression results
        rowNumber = 2
        for item in lm_corr.index.to_list():
            ws.cell(row=rowNumber, column=1, value=item)
            ws.cell(row=rowNumber, column=2, value=lm_corr['m'][item])
            ws.cell(row=rowNumber, column=3, value=lm_corr['c'][item])
            ws.cell(row=rowNumber, column=4, value=lm_corr['rsquared'][item])
            ws.cell(row=rowNumber, column=5, value=lm_corr['difference'][item])
            ws.cell(row=rowNumber, column=6, value=lm_corr['mse'][item])
            ws.cell(row=rowNumber, column=7, value=lm_corr['rmse'][item])
            rowNumber = rowNumber + 1

        if stabilityFlag:
            rowNumber = 2
            className = 1
            for i in range(0,len(lm_corr_stability)):
                start = className *8 + 1
                try:
                    for item in lm_corr_stability[i].index.to_list():
                        ws.cell(row = rowNumber, column = start, value = item)
                        ws.cell(row=rowNumber, column=start+1, value=lm_corr_stability[i]['m'][item])
                        ws.cell(row=rowNumber, column=start+2, value=lm_corr_stability[i]['c'][item])
                        ws.cell(row=rowNumber, column=start+3, value=lm_corr_stability[i]['rsquared'][item])
                        ws.cell(row=rowNumber, column=start+4, value=lm_corr_stability[i]['difference'][item])
                        ws.cell(row=rowNumber, column=start+5, value=lm_corr_stability[i]['mse'][item])
                        ws.cell(row=rowNumber, column=start+6, value=lm_corr_stability[i]['rmse'][item])
                        rowNumber = rowNumber + 1
                except:
                    pass
                className = className + 1
                rowNumber = 2

        if cup_alphaFlag:
            rowNumber = 14
            className = 1
            for i in range(0,len(lm_corr_stability_alpha_Ane)):
                start = className *8 + 1
                try:
                    for item in lm_corr_stability_alpha_Ane[i].index.to_list():
                        ws.cell(row = rowNumber, column = start, value = item)
                        ws.cell(row=rowNumber, column=start+1, value=lm_corr_stability_alpha_Ane[i]['m'][item])
                        ws.cell(row=rowNumber, column=start+2, value=lm_corr_stability_alpha_Ane[i]['c'][item])
                        ws.cell(row=rowNumber, column=start+3, value=lm_corr_stability_alpha_Ane[i]['rsquared'][item])
                        ws.cell(row=rowNumber, column=start+4, value=lm_corr_stability_alpha_Ane[i]['difference'][item])
                        ws.cell(row=rowNumber, column=start+5, value=lm_corr_stability_alpha_Ane[i]['mse'][item])
                        ws.cell(row=rowNumber, column=start+6, value=lm_corr_stability_alpha_Ane[i]['rmse'][item])
                        rowNumber = rowNumber + 1
                except:
                    pass
                className = className + 1
                rowNumber = 14

        if RSD_alphaFlag:
            rowNumber = 26
            className = 1
            for i in range(0,len(lm_corr_stability_alpha_RSD)):
                start = className *8 + 1
                try:
                    for item in lm_corr_stability_alpha_RSD[i].index.to_list():
                        ws.cell(row = rowNumber, column = start, value = item)
                        ws.cell(row=rowNumber, column=start+1, value=lm_corr_stability_alpha_RSD[i]['m'][item])
                        ws.cell(row=rowNumber, column=start+2, value=lm_corr_stability_alpha_RSD[i]['c'][item])
                        ws.cell(row=rowNumber, column=start+3, value=lm_corr_stability_alpha_RSD[i]['rsquared'][item])
                        ws.cell(row=rowNumber, column=start+4, value=lm_corr_stability_alpha_RSD[i]['difference'][item])
                        ws.cell(row=rowNumber, column=start+5, value=lm_corr_stability_alpha_RSD[i]['mse'][item])
                        ws.cell(row=rowNumber, column=start+6, value=lm_corr_stability_alpha_RSD[i]['rmse'][item])
                        rowNumber = rowNumber + 1
                except:
                    pass
                className = className + 1
                rowNumber = 26

        rowNumber = 37

        Result_df_mean_1mps = pd.DataFrame()
        Result_df_mean_05mps = pd.DataFrame()
        Result_df_std_1mps = pd.DataFrame()
        Result_df_std_05mps = pd.DataFrame()
        Result_df_mean_tiRef = pd.DataFrame()
        Result_df_std_tiRef = pd.DataFrame()

        classes = ['class1', 'class2', 'class3', 'class4', 'class5']

        # --- All data
        # TI values by ws bin
        Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps,Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TIbybin, Result_df_mean_1mps,Result_df_mean_05mps,Result_df_std_1mps,Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef)
        # TI MBE
        Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_MBE_j_, Result_df_mean_1mps,Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef)
        # RepTI MBE
        Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_MBE_j_, Result_df_mean_1mps,Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef)
        # TI RMSE
        if TI_RMSE_j_ is not None:
            for val in TI_RMSE_j_:
                for i in val:
                    dat = i['mean']
                    new_data = dat.add_prefix(str('mean' + '_'))
                    if new_data.columns.name == 'bins':
                        Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                    elif new_data.columns.name == 'bins_p5':
                        Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])
        # RepTI RMSE
        if RepTI_RMSE_j_ is not None:
            for val in RepTI_RMSE_j_:
                for i in val:
                    dat = i['mean']
                    new_data = dat.add_prefix(str('mean' + '_'))
                    if new_data.columns.name == 'bins':
                        Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                    elif new_data.columns.name == 'bins_p5':
                        Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])

        # TI DIff
        Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_Diff_j_,Result_df_mean_1mps, Result_df_mean_05mps,Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef)
        # RepTI DIff
        Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_Diff_j_,Result_df_mean_1mps,Result_df_mean_05mps, Result_df_std_1mps,Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef)

        # --- Stability Classes
        if stabilityFlag:
            for i in range(0,len(TIbybin_stability)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TIbybin_stability[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(TI_MBE_j_stability)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_MBE_j_stability[i], Result_df_mean_1mps, Result_df_mean_05mps,Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(RepTI_MBE_j_stability)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_MBE_j_stability[i], Result_df_mean_1mps, Result_df_mean_05mps,Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(TI_Diff_j_stability)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_Diff_j_stability[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(RepTI_Diff_j_stability)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_Diff_j_stability[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for j in range(0,len(TI_RMSE_j_stability)):
                stabilityClass = classes[j]
                df_list = TI_RMSE_j_stability[j]
                if isinstance(df_list,list) == False:
                   pass
                else:
                   for val in df_list:
                       for i in val:
                           dat = i['mean']
                           new_data = dat.add_prefix(str('mean' + '_'))
                           new_index = [str('stability_' + stabilityClass + '_' + n) for n in new_data.index.to_list()]
                           new_data.index = new_index
                           if new_data.columns.name == 'bins':
                              Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                           elif new_data.columns.name == 'bins_p5':
                              Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])
            for j in range(0,len(RepTI_RMSE_j_stability)):
                stabilityClass = classes[j]
                df_list = RepTI_RMSE_j_stability[j]
                if isinstance(df_list,list) == False:
                    pass
                else:
                    for val in df_list:
                        for i in val:
                            dat = i['mean']
                            new_data = dat.add_prefix(str('mean' + '_'))
                            new_index = [str('stability_' + stabilityClass + '_' + n) for n in new_data.index.to_list()]
                            new_data.index = new_index
                            if new_data.columns.name == 'bins':
                               Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                            elif new_data.columns.name == 'bins_p5':
                               Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])

        # --- stability class by alpha Ane
        if cup_alphaFlag:
            pass
        else:
            TIbybin_stability_alpha_Ane = TIbybin_stability
            TI_MBE_j_stability_alpha_Ane = TI_MBE_j_stability
            RepTI_MBE_j_stability_alpha_Ane = RepTI_MBE_j_stability
            RepTI_Diff_j_stability_alpha_Ane = RepTI_Diff_j_stability
            TI_Diff_j_stability_alpha_Ane = TI_Diff_j_stability
            TI_RMSE_j_stability_alpha_Ane = TI_RMSE_j_stability
            RepTI_RMSE_j_stability_alpha_Ane = RepTI_RMSE_j_stability
        if stabilityFlag:
            for i in range(0,len(TIbybin_stability_alpha_Ane)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TIbybin_stability_alpha_Ane[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(TI_MBE_j_stability_alpha_Ane)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_MBE_j_stability_alpha_Ane[i], Result_df_mean_1mps, Result_df_mean_05mps,Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(RepTI_MBE_j_stability_alpha_Ane)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_MBE_j_stability_alpha_Ane[i], Result_df_mean_1mps, Result_df_mean_05mps,Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(TI_Diff_j_stability_alpha_Ane)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_Diff_j_stability_alpha_Ane[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(RepTI_Diff_j_stability_alpha_Ane)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_Diff_j_stability_alpha_Ane[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for j in range(0,len(TI_RMSE_j_stability_alpha_Ane)):
                stabilityClass = classes[j]
                df_list = TI_RMSE_j_stability_alpha_Ane[j]
                if isinstance(df_list,list) == False:
                   pass
                else:
                   for val in df_list:
                       for i in val:
                           dat = i['mean']
                           new_data = dat.add_prefix(str('mean' + '_'))
                           new_index = [str('stability_' + stabilityClass + '_' + n) for n in new_data.index.to_list()]
                           new_data.index = new_index
                           if new_data.columns.name == 'bins':
                              Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                           elif new_data.columns.name == 'bins_p5':
                              Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])
            for j in range(0,len(RepTI_RMSE_j_stability_alpha_Ane)):
                stabilityClass = classes[j]
                df_list = RepTI_RMSE_j_stability_alpha_Ane[j]
                if isinstance(df_list,list) == False:
                    pass
                else:
                    for val in df_list:
                        for i in val:
                            dat = i['mean']
                            new_data = dat.add_prefix(str('mean' + '_'))
                            new_index = [str('stability_' + stabilityClass + '_' + n) for n in new_data.index.to_list()]
                            new_data.index = new_index
                            if new_data.columns.name == 'bins':
                               Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                            elif new_data.columns.name == 'bins_p5':
                               Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])

        # --- stability class by alpha RSD
        if RSD_alphaFlag:
            pass
        else:
            TIbybin_stability_alpha_RSD = TIbybin_stability
            TI_MBE_j_stability_alpha_RSD = TI_MBE_j_stability
            RepTI_MBE_j_stability_alpha_RSD = RepTI_MBE_j_stability
            RepTI_Diff_j_stability_alpha_RSD = RepTI_Diff_j_stability
            TI_Diff_j_stability_alpha_RSD = TI_Diff_j_stability
            TI_RMSE_j_stability_alpha_RSD = TI_RMSE_j_stability
            RepTI_RMSE_j_stability_alpha_RSD = RepTI_RMSE_j_stability
            for i in range(0,len(TIbybin_stability_alpha_RSD)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TIbybin_stability_alpha_RSD[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(TI_MBE_j_stability_alpha_RSD)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_MBE_j_stability_alpha_RSD[i], Result_df_mean_1mps, Result_df_mean_05mps,Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(RepTI_MBE_j_stability_alpha_RSD)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_MBE_j_stability_alpha_RSD[i], Result_df_mean_1mps, Result_df_mean_05mps,Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(TI_Diff_j_stability_alpha_RSD)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_Diff_j_stability_alpha_RSD[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for i in range(0,len(RepTI_Diff_j_stability_alpha_RSD)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_Diff_j_stability_alpha_RSD[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i])
            for j in range(0,len(TI_RMSE_j_stability_alpha_RSD)):
                stabilityClass = classes[j]
                df_list = TI_RMSE_j_stability_alpha_RSD[j]
                if isinstance(df_list,list) == False:
                   pass
                else:
                   for val in df_list:
                       for i in val:
                           dat = i['mean']
                           new_data = dat.add_prefix(str('mean' + '_'))
                           new_index = [str('stability_' + stabilityClass + '_' + n) for n in new_data.index.to_list()]
                           new_data.index = new_index
                           if new_data.columns.name == 'bins':
                              Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                           elif new_data.columns.name == 'bins_p5':
                              Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])
            for j in range(0,len(RepTI_RMSE_j_stability_alpha_RSD)):
                stabilityClass = classes[j]
                df_list = RepTI_RMSE_j_stability_alpha_RSD[j]
                if isinstance(df_list,list) == False:
                    pass
                else:
                    for val in df_list:
                        for i in val:
                            dat = i['mean']
                            new_data = dat.add_prefix(str('mean' + '_'))
                            new_index = [str('stability_' + stabilityClass + '_' + n) for n in new_data.index.to_list()]
                            new_data.index = new_index
                            if new_data.columns.name == 'bins':
                               Result_df_mean_1mps = pd.concat([Result_df_mean_1mps,new_data])
                            elif new_data.columns.name == 'bins_p5':
                               Result_df_mean_05mps = pd.concat([Result_df_mean_05mps,new_data])

        write_resultstofile(Result_df_mean_1mps, ws, rowNumber,1)
        rowNumber += len(Result_df_mean_1mps) + 3
        write_resultstofile(Result_df_mean_05mps, ws, rowNumber,1)
        rowNumber += len(Result_df_mean_05mps) + 3
        write_resultstofile(Result_df_std_1mps, ws, rowNumber,1)
        rowNumber += len(Result_df_std_1mps) + 3
        write_resultstofile(Result_df_std_05mps, ws, rowNumber,1)
        rowNumber += len(Result_df_std_05mps) + 3

        # --- All data
        # TI values by Ref TI bin
        Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TIbyRefbin,Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, byTIRefbin = True)
        if stabilityFlag:
            for i in range(0,len(TIbyRefbin_stability)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TIbyRefbin_stability[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)
        if cup_alphaFlag:
            for i in range(0,len(TIbyRefbin_stability_alpha_Ane)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TIbyRefbin_stability_alpha_Ane[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)
        if RSD_alphaFlag:
            for i in range(0,len(TIbyRefbin_stability_alpha_RSD)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TIbyRefbin_stability_alpha_RSD[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)

        # TI Diff by Ref TI bin
        Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_Diff_r_,Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, byTIRefbin = True)
        Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_Diff_r_,Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, byTIRefbin = True)
        if stabilityFlag:
            for i in range(0,len(TI_Diff_r_stability)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_Diff_r_stability[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_Diff_r_stability[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)
        if cup_alphaFlag:
            for i in range(0,len(TI_Diff_r_stability_alpha_Ane)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_Diff_r_stability_alpha_Ane[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_Diff_r_stability_alpha_Ane[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)
        if RSD_alphaFlag:
            for i in range(0,len(TI_Diff_r_stability_alpha_RSD)):
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(TI_Diff_r_stability_alpha_RSD[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)
                Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps,Result_df_mean_tiRef, Result_df_std_tiRef = configure_for_printing(RepTI_Diff_r_stability_alpha_RSD[i], Result_df_mean_1mps, Result_df_mean_05mps, Result_df_std_1mps, Result_df_std_05mps, Result_df_mean_tiRef, Result_df_std_tiRef, stabilityClass=classes[i], byTIRefbin = True)

        write_resultstofile(Result_df_mean_tiRef, ws, rowNumber,1)
        rowNumber += len(Result_df_mean_tiRef) + 3
        write_resultstofile(Result_df_std_tiRef, ws, rowNumber,1)
        rowNumber += len(Result_df_std_tiRef) + 3

        # total stats
        if isinstance(total_stats, pd.DataFrame) and isinstance(total_stats, tuple) == False:
            write_resultstofile(total_stats, ws, rowNumber, 1)
            rowNumber += len(total_stats) + 3
        else:
            pass
        # total stats by stability
        if stabilityFlag:
            for i in range(0,len(total_stats_stability)):
                if isinstance(total_stats_stability[i],pd.DataFrame) and isinstance(total_stats_stability[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('Total stats by stability_' + 'Class_1'))
                    rowNumber += 1
                    write_resultstofile(total_stats_stability[i], ws, rowNumber, 1)
                    rowNumber += len(total_stats_stability[i]) + 3
                else:
                    pass
        if cup_alphaFlag:
            for i in range(0,len(total_stats_stability_alpha_Ane)):
                if isinstance(total_stats_stability_alpha_Ane[i],pd.DataFrame) and isinstance(total_stats_stability_alpha_Ane[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('Total stats by stability_alpha_Ane_' + 'Class_1'))
                    rowNumber += 1
                    write_resultstofile(total_stats_stability_alpha_Ane[i], ws, rowNumber, 1)
                    rowNumber += len(total_stats_stability_alpha_Ane[i]) + 3
                else:
                    pass
        if RSD_alphaFlag:
            for i in range(0,len(total_stats_stability_alpha_RSD)):
                if isinstance(total_stats_stability_alpha_RSD[i],pd.DataFrame) and isinstance(total_stats_stability_alpha_RSD[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('Total stats by stability_alpha_RSD_' + 'Class_1'))
                    rowNumber += 1
                    write_resultstofile(total_stats_stability_alpha_RSD[i], ws, rowNumber, 1)
                    rowNumber += len(total_stats_stability_alpha_RSD[i]) + 3
                else:
                    pass

        # representative stats: belownominal
        if isinstance(belownominal_stats, pd.DataFrame) and isinstance(belownominal_stats, tuple) == False:
            write_resultstofile(belownominal_stats, ws, rowNumber, 1)
            rowNumber += len(belownominal_stats) + 3
        else:
            pass
        # representative by stability: below nominal
        if stabilityFlag:
            for i in range(0,len(belownominal_stats_stability)):
                if isinstance(belownominal_stats_stability[i],pd.DataFrame) and isinstance(belownominal_stats_stability[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('belownominal stats by stability_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    write_resultstofile(belownominal_stats_stability[i], ws, rowNumber, 1)
                    rowNumber += len(belownominal_stats_stability[i]) + 3
                else:
                    pass
        if cup_alphaFlag:
            for i in range(0,len(belownominal_stats_stability_alpha_Ane)):
                if isinstance(belownominal_stats_stability_alpha_Ane[i],pd.DataFrame) and isinstance(belownominal_stats_stability_alpha_Ane[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('belownominal stats by stability_alpha_Ane_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    write_resultstofile(belownominal_stats_stability_alpha_Ane[i], ws, rowNumber, 1)
                    rowNumber += len(belownominal_stats_stability_alpha_Ane[i]) + 3
                else:
                    pass
        if RSD_alphaFlag:
            for i in range(0,len(belownominal_stats_stability_alpha_RSD)):
                if isinstance(belownominal_stats_stability_alpha_RSD[i],pd.DataFrame) and isinstance(belownominal_stats_stability_alpha_RSD[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('belownominal stats by stability_alpha_RSD_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    write_resultstofile(belownominal_stats_stability_alpha_RSD[i], ws, rowNumber, 1)
                    rowNumber += len(belownominal_stats_stability_alpha_RSD[i]) + 3
                else:
                    pass

        # representative stats: abovenominal
        if isinstance(abovenominal_stats, pd.DataFrame) and isinstance(abovenominal_stats, tuple) == False:
            write_resultstofile(abovenominal_stats, ws, rowNumber, 1)
            rowNumber += len(abovenominal_stats) + 3
        else:
            pass
        # representative by stability: below nominal
        if stabilityFlag:
            for i in range(0,len(abovenominal_stats_stability)):
                if isinstance(abovenominal_stats_stability[i],pd.DataFrame) and isinstance(abovenominal_stats_stability[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('abovenominal stats by stability_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    write_resultstofile(abovenominal_stats_stability[i], ws, rowNumber, 1)
                    rowNumber += len(abovenominal_stats_stability[i]) + 3
                else:
                    pass
        if cup_alphaFlag:
            for i in range(0,len(abovenominal_stats_stability_alpha_Ane)):
                if isinstance(abovenominal_stats_stability_alpha_Ane[i],pd.DataFrame) and isinstance(abovenominal_stats_stability_alpha_Ane[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('abovenominal stats by stability_alpha_Ane_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    write_resultstofile(abovenominal_stats_stability_alpha_Ane[i], ws, rowNumber, 1)
                    rowNumber += len(abovenominal_stats_stability_alpha_Ane[i]) + 3
                else:
                    pass
        if RSD_alphaFlag:
            for i in range(0,len(abovenominal_stats_stability_alpha_RSD)):
                if isinstance(abovenominal_stats_stability_alpha_RSD[i],pd.DataFrame) and isinstance(abovenominal_stats_stability_alpha_RSD[i], tuple) == False:
                    ws.cell(row=rowNumber, column=1, value= str('abovenominal stats by stability_alpha_RSD_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    write_resultstofile(abovenominal_stats_stability_alpha_RSD[i], ws, rowNumber, 1)
                    rowNumber += len(abovenominal_stats_stability_alpha_RSD[i]) + 3
                else:
                    pass

        # Representative TI
        if isinstance(rep_TI_results_1mps, pd.DataFrame):
            write_resultstofile(rep_TI_results_1mps, ws, rowNumber,1)
            rowNumber += len(rep_TI_results_1mps) + 3
        else:
            pass
        if stabilityFlag:
            for i in range(0,len(rep_TI_results_1mps_stability)):
                if isinstance(rep_TI_results_1mps_stability[i], pd.DataFrame):
                    ws.cell(row=rowNumber, column=1, value= str('representative stats by stability_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    try:
                        write_resultstofile(rep_TI_results_1mps_stability[i], ws, rowNumber, 1)
                        rowNumber += len(rep_TI_results_1mps_stability[i]) + 3
                    except:
                        ws.cell(row=rowNumber, column=1, value= np.NaN)
                        rowNumber += 4
                else:
                    pass
        if cup_alphaFlag:
            for i in range(0,len(rep_TI_results_1mps_stability_alpha_Ane)):
                if isinstance(rep_TI_results_1mps_stability_alpha_Ane[i], pd.DataFrame):
                    ws.cell(row=rowNumber, column=1, value= str('representative stats by stability_alpha_Ane_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    try:
                        write_resultstofile(rep_TI_results_1mps_stability_alpha_RSD[i], ws, rowNumber, 1)
                        rowNumber += len(rep_TI_results_1mps_stability_alpha_RSD[i]) + 3
                    except:
                        ws.cell(row=rowNumber, column=1, value= np.NaN)
                        rowNumber += 4
                else:
                    pass
        if RSD_alphaFlag:
            for i in range(0,len(rep_TI_results_1mps_stability_alpha_RSD)):
                if isinstance(rep_TI_results_1mps_stability_alpha_RSD[i], pd.DataFrame):
                    ws.cell(row=rowNumber, column=1, value= str('representative stats by stability_alpha_RSD_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    try:
                        write_resultstofile(rep_TI_results_1mps_stability_alpha_RSD[i], ws, rowNumber, 1)
                        rowNumber += len(rep_TI_results_1mps_stability_alpha_RSD[i]) + 3
                    except:
                        ws.cell(row=rowNumber, column=1, value= np.NaN)
                        rowNumber += 4
                else:
                    pass

        if isinstance(rep_TI_results_05mps, pd.DataFrame):
            write_resultstofile(rep_TI_results_05mps, ws, rowNumber,1)
            rowNumber += len(rep_TI_results_05mps) + 3
        else:
            pass
        if stabilityFlag:
            for i in range(0,len(rep_TI_results_05mps_stability)):
                if isinstance(rep_TI_results_05mps_stability[i], pd.DataFrame):
                    ws.cell(row=rowNumber, column=1, value= str('representative stats by stability_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    write_resultstofile(rep_TI_results_05mps_stability[i], ws, rowNumber, 1)
                    rowNumber += len(rep_TI_results_05mps_stability[i]) + 3
                else:
                    pass
        if cup_alphaFlag:
            for i in range(0,len(rep_TI_results_05mps_stability_alpha_Ane)):
                if isinstance(rep_TI_results_05mps_stability[i], pd.DataFrame):
                    ws.cell(row=rowNumber, column=1, value= str('representative stats by stability_alpha_Ane_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    try:
                        write_resultstofile(rep_TI_results_05mps_stability_alpha_Ane[i], ws, rowNumber, 1)
                        rowNumber += len(rep_TI_results_05mps_stability_alpha_Ane[i]) + 3
                    except:
                        ws.cell(row=rowNumber, column=1, value= np.NaN)
                        rowNumber += 4
                else:
                    pass
        if RSD_alphaFlag:
            for i in range(0,len(rep_TI_results_05mps_stability_alpha_RSD)):
                if isinstance(rep_TI_results_05mps_stability_alpha_RSD[i], pd.DataFrame):
                    ws.cell(row=rowNumber, column=1, value= str('representative stats by stability_alpha_RSD_' + 'Class_' + str(i+1)))
                    rowNumber += 1
                    try:
                        write_resultstofile(rep_TI_results_05mps_stability_alpha_RSD[i], ws, rowNumber, 1)
                        rowNumber += len(rep_TI_results_05mps_stability_alpha_RSD[i]) + 3
                    except:
                        ws.cell(row=rowNumber, column=1, value= np.NaN)
                        rowNumber += 4
                else:
                    pass

    d = wb.create_sheet(title='K-S Tests')
    for r in dataframe_to_rows(Dist_stats_df, index=False):
        d.append(r)
    f = wb.create_sheet(title='Extrapolation Metadata')
    for r in dataframe_to_rows(extrap_metadata, index=False):
       f.append(r)

    # remove blank sheet
    del wb['Sheet']

    wb.save(results_filename)
